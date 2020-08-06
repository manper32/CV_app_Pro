from django.shortcuts import render
from django.http import HttpResponse
from datetime import datetime
import psycopg2
import math
import pandas as pd
from openpyxl import Workbook
import csv

def psql_pdc(query):
    #credenciales PostgreSQL produccion
    connP_P = {
    	'host' : '10.150.1.74',
    	'port' : '5432',
    	'user':'postgres',
    	'password':'cobrando.bi.2020',
    	'database' : 'postgres'}

    #conexion a PostgreSQL produccion
    conexionP_P = psycopg2.connect(**connP_P)
    #print('\nConexión con el servidor PostgreSQL produccion establecida!')
    cursorP_P = conexionP_P.cursor ()
    #ejecucion query telefonos PostgreSQL
    cursorP_P.execute(query)
    anwr = cursorP_P.fetchall()
    
    cursorP_P.close()
    conexionP_P.close()
    return anwr
    
def to_horiz(anwr_P,name,_id):
    #vertical horizontal 
    anwr_P1 = anwr_P.pivot(index=0,columns=1)
    anwr_P1[_id] = anwr_P1.index
    
    col1 = []
    i=0
    for i in range(anwr_P1.shape[1]-1):
        col1.append(name+str(i+1))
    col1.append(_id)
    
    anwr_P1.columns = col1
    
    return anwr_P1

def csv_o(fn,name):
    response = HttpResponse(content_type = "text/csv")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content

    # for j in range(fn.shape[1]):
    #     try:
    #         fn.iloc[:,j] = fn.iloc[:,j].str.decode(encoding='utf-8-sig')
    #         fn.iloc[:,j] = fn.iloc[:,j].str.encode(encoding='utf_16_le')
    #     except:
    #         pass

    fn2 = [tuple(x) for x in fn.values]
    writer = csv.writer(response,delimiter ='|')
    writer.writerow(fn.columns)
    writer.writerows(fn2)

    return response

def excel(fn,name):
    wb = Workbook()
    ws = wb.active

    k = 0
    a = pd.DataFrame(fn.columns)

    for k in range(a.shape[0]):
        ws.cell(row = 1, column = k+1).value = a.iloc[k,0]

    i=0
    j=0

    for i in range(fn.shape[0]):
        for j in range(0,fn.shape[1]):
            try:
                ws.cell(row = i+2, column = j+1).value = fn.iloc[i,j]
            except:
                pass
            
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = %s"%name
    response["Content-Disposition"] = content
    wb.save(response)
    return response

def excel_CV_COL(request):
    today = datetime.now()
    tablename = "CV_Col"+today.strftime("%Y%m%d%H")+'.xlsx'

    with open("./hello/Plantillas/Colp/QueryTel_COL.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Colp/QueryCor_COL.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Colp/QueryDir_COL.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Colp/QueryCV_COL.txt","r") as f4:
        queryP_cons = f4.read()

    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")


    df = df.rename(columns={0:'rownumber',
                            1:'obligacion_id',
                            2:'deudor_id',
                            3:'unico',
                            4:'estado',
                            5:'tipo_cliente',
                            6:'nombre',
                            7:'producto',
                            8:'initial_bucket',
                            9:'ciudad',
                            10:'sucursal',
                            11:'tipo_prod',
                            12:'dias_mora_inicial',
                            13:'dias_mora_actual',
                            14:'rango_mora_inicial',
                            15:'rango_mora_final',
                            16:'rango',
                            17:'suma_pareto',
                            18:'rango_pareto',
                            19:'fcast',
                            20:'fdesem',
                            21:'vrdesem',
                            22:'saldo_total_inicial',
                            23:'saldo_total_actual',
                            24:'saldo_capital_inicial',
                            25:'saldo_capital_actual',
                            26:'saldo_vencido_inicial',
                            27:'saldo_vencido_actual',
                            28:'pagomin',
                            29:'fultpago',
                            30:'vrultpago',
                            31:'agencia',
                            32:'tasainter',
                            33:'feultref',
                            34:'ultcond',
                            35:'fasigna',
                            36:'eqasicampana',
                            37:'diferencia_pago',
                            38:'pago_preliminar',
                            39:'pago_cliente',
                            40:'min',
                            41:'tarifa',
                            42:'honorarios',
                            43:'perfil_mes_4',
                            44:'perfil_mes_3',
                            45:'perfil_mes_2',
                            46:'perfil_mes_1',
                            47:'fecha_primer_gestion',
                            48:'fecha_ultima_gestion',
                            49:'perfil_mes_actual',
                            50:'contactabilidad',
                            51:'ultimo_alo',
                            52:'descod1',
                            53:'descod2',
                            54:'asesor',
                            55:'fecha_gestion',
                            56:'telefono_mejor_gestion',
                            57:'mejorgestionhoy',
                            58:'repeticion',
                            59:'llamadas',
                            60:'sms',
                            61:'correos',
                            62:'gescall',
                            63:'visitas',
                            64:'whatsapp',
                            65:'no_contacto',
                            66:'total_gestiones',
                            67:'telefono_positivo',
                            68:'marcaciones_telefono_positivo',
                            69:'ultima_marcacion_telefono_positivo',
                            70:'fec_creacion_ult_compromiso',
                            71:'fec_pactada_ult_compromiso',
                            72:'valor_acordado_ult_compromiso',
                            73:'asesor_ult_compromiso',
                            74:'cantidad_acuerdos_mes',
                            75:'estado_acuerdo'
                            })

    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return excel(fn,tablename)


def csv_CV_Claro(request):
    today = datetime.now()
    tablename = "CV_Claro"+today.strftime("%Y%m%d%H")+'.csv'

    with open("./hello/Plantillas/Claro/QueryTel_Claro.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Claro/QueryCor_Claro.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Claro/QueryCV_Claro.txt","r") as f4:
        queryP_cons = f4.read()
        
    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    yanwr = psql_pdc(queryP_cons)    

    #dataframes
    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    df = pd.DataFrame(yanwr)

    anwr_P1 = to_horiz(anwr_P,'phone',"deudor_id")

    #renombrar campos correos
    anwr_C = anwr_C.rename(columns={
                            
                            0:'deudor_id',
                            1:'mail0',
                            2:'mail1'})

    #renombrar campos CV
    df = df.rename(columns={0:'rownumber',
                            1:'deudor_id',
                            2:'obligacion_id',
                            3:'nombredelcliente',
                            4:'estado',
                            5:'tipo_cliente',
                            6:'unico',
                            7:'crmorigen',
                            8:'potencialmark',
                            9:'prepotencialmark',
                            10:'writeoffmark',
                            11:'segmento_bpo',
                            12:'valorscoring',
                            13:'numeroreferenciadepago',
                            14:'monto_inicial',
                            15:'monto_ini_cuenta',
                            16:'porcentaje_descuento',
                            17:'val_descuento',
                            18:'val_a_pagar',
                            19:'deuda_real',
                            20:'val_pago',
                            21:'maxfec_pago',
                            22:'mfecha_compromiso',
                            23:'fecha_pago',
                            24:'valor_compromiso',
                            25:'estado_acuerdo',
                            26:'indicador_m4',
                            27:'indicador_m3',
                            28:'indicador_m2',
                            29:'indicador_m1',
                            30:'fecha_primer_gestion',
                            31:'fecha_ultima_gestion',
                            32:'indicador_mes_actual',
                            33:'tel_mes_actual',
                            34:'asesor_mes_actual',
                            35:'fec_indicador_mes_actual',
                            36:'contact',
                            37:'indicador_hoy',
                            38:'repeticion_mes_actual',
                            39:'llamadas_mes_actual',
                            40:'sms_mes_actual',
                            41:'correos_mes_actual',
                            42:'gescall_mes_actual',
                            43:'whatsapp_mes_actual',
                            44:'visitas',
                            45:'no_contacto_mes_actual',
                            46:'total_ges_mes_actual',
                            47:'tel_positivo',
                            48:'ult_fec_tel_pos'})

    i=0
    lin = ['no_contacto_mes_actual','gescall_mes_actual','tel_mes_actual','tel_positivo']
    for i in lin:
        df[i].fillna(0,inplace=True)
        df[i] = df[i].apply(lambda x: round(x))
        df[i] = df[i].astype('str')

    fn = pd.merge(df,anwr_P1,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,anwr_C,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return csv_o(fn,tablename)

def csv_CV_CarP(request):
    today = datetime.now()
    tablename = "CV_CarP"+today.strftime("%Y%m%d%H")+'.csv'

    with open("./hello/Plantillas/CarP/QueryTel_CarP.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/CarP/QueryCor_CarP.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/CarP/QueryDir_CarP.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/CarP/QueryCV_CarP.txt","r") as f4:
        queryP_cons = f4.read()


    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")

    #renombrar campos CV
    df = df.rename(columns={0:'deudor_id',
                            1:'unico',
                            2:'nombre',
                            3:'obligacion',
                            4:'obligacion_17',
                            5:'tipo_cliente',
                            6:'sucursal_final',
                            7:'zona',
                            8:'ano_castigo',
                            9:'saldo_k_pareto_mes_vigente',
                            10:'intereses',
                            11:'honorarios_20',
                            12:'saldo_total_mes_vigente',
                            13:'saldo_total_pareto_mes_vigente_',
                            14:'saldokpareto',
                            15:'rango_k_pareto',
                            16:'interesespareto',
                            17:'honorariospareto',
                            18:'porcentaje_k_del_total',
                            19:'porcentaje_intereses_del_total',
                            20:'porcentaje_honorarios_del_total',
                            21:'rango_k_porcentaje',
                            22:'capital_20_porciento',
                            23:'dias_mora_acumulado',
                            24:'marca_juridica_cliente',
                            25:'investigacion_de_bienes',
                            26:'valor_pago',
                            27:'ultima_fecha',
                            28:'valor_compromiso',
                            29:'fecha_compromiso',
                            30:'fecha_pactada_compromiso',
                            31:'asesor',
                            32:'estado_acuerdo',
                            33:'ind_m4',
                            34:'ind_m3',
                            35:'ind_m2',
                            36:'ind_m1',
                            37:'fecha_primer_gestion',
                            38:'fecha_ultima_gestion',
                            39:'indicador',
                            40:'telefono_mejor_gestion',
                            41:'asesor',
                            42:'fecha_gestion',
                            43:'contactabilidad',
                            44:'indicador_hoy',
                            45:'repeticion',
                            46:'llamadas',
                            47:'sms',
                            48:'correos',
                            49:'gescall',
                            50:'whatsapp',
                            51:'visitas',
                            52:'no_contacto',
                            53:'total_gestiones',
                            54:'telefono_positivo',
                            55:'fec_ultima_marcacion'})


    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)

    return csv_o(fn,tablename)                

def csv_CV_Fala(request):
    today = datetime.now()
    tablename = "CV_Fal"+today.strftime("%Y%m%d%H")+'.csv'

    with open("./hello/Plantillas/Fala/QueryTel_Fal.txt","r") as f1:
        queryP_PT = f1.read()
        
    with open("./hello/Plantillas/Fala/QueryCor_Fal.txt","r") as f2:
        queryP_PC = f2.read()
            
    with open("./hello/Plantillas/Fala/QueryDir_Fal.txt","r") as f3:
        queryP_PD = f3.read()
            
    with open("./hello/Plantillas/Fala/QueryCV_Fal.txt","r") as f4:
        queryP_cons = f4.read()

    with open("./hello/Plantillas/Fala/QueryPa_Fal.txt","r") as f5:
        queryP_Pa = f5.read()

    with open("./hello/Plantillas/Fala/QueryRe_Fal.txt","r") as f6:
        queryP_PR = f6.read()

    anwr = psql_pdc(queryP_PT)
    anwrC = psql_pdc(queryP_PC)
    anwrD = psql_pdc(queryP_PD)
    anwrPa = psql_pdc(queryP_Pa)
    anwrR = psql_pdc(queryP_PR)
    yanwr = psql_pdc(queryP_cons)

    anwr_P = pd.DataFrame(anwr)
    anwr_C = pd.DataFrame(anwrC)
    anwr_D = pd.DataFrame(anwrD)
    anwr_Pa = pd.DataFrame(anwrPa)
    anwr_R = pd.DataFrame(anwrR)
    df = pd.DataFrame(yanwr)

    inf = to_horiz(anwr_P,'phone',"deudor_id")
    infC = to_horiz(anwr_C,'mail',"deudor_id")
    infD = to_horiz(anwr_D,'address',"deudor_id")
    infR = to_horiz(anwr_R,'referencia',"deudor_id")
    try:
        infP = to_horiz(anwr_Pa,'pago','obligacion_id')
    except:
        pass

    #renombrar campos CV
    df = df.rename(columns={0:'tipo_producto_asignacion',
                            1:'grupo',
                            2:'cartera',
                            3:'tipo_cliente',
                            4:'unico',
                            5:'unico_pro',
                            6:'obligacion_id',
                            7:'deudor_id',
                            8:'nombre',
                            9:'producto',
                            10:'saldototal',
                            11:'saldo_pareto',
                            12:'segmentacion',
                            13:'peso',
                            14:'alturamora_hoy',
                            15:'rango',
                            16:'dias_mora',
                            17:'vencto',
                            18:'mejor_gestion',
                            19:'total_gestiones',
                            20:'fecha_ultima_gestion',
                            21:'asesor_mejor_gestion',
                            22:'fecha_compromiso',
                            23:'fecha_pago_compromiso',
                            24:'valor_compromiso',
                            25:'estado_acuerdo',
                            26:'dias_mora_pagos',
                            27:'valor_pago',
                            28:'fecha_pago',
                            29:'pagos_reales',
                            30:'pago_para_honorarios',
                            31:'pago_para_factura',
                            32:'pendiente',
                            33:'pago_total',
                            34:'nvo_status',
                            35:'status_refresque',
                            36:'nvo_status_refresque',
                            37:'dias_mora_refresque',
                            38:'pendiente_mas_gastos',
                            39:'vencida_mas_gastos',
                            40:'gastos_mora',
                            41:'gastos_cv',
                            42:'porcentaje_gasto',
                            43:'valor_a_mantener_sin_gxc',
                            44:'cv8',
                            45:'cv9',
                            46:'cv10',
                            47:'cv11',
                            48:'cv12',
                            49:'restructuracion',
                            50:'valor',
                            51:'pagominimo_anterior',
                            52:'pagominimo_actual',
                            53:'cuota36',
                            54:'cuota42',
                            55:'cuota60',
                            56:'cuota72',
                            57:'proyectada_cargue',
                            58:'aplica_ajuste',
                            59:'fecha',
                            60:'diferencia',
                            61:'porcentaje_saldo_total',
                            62:'x',
                            63:'valor',
                            64:'porcentaje_participacion',
                            65:'ind_m4',
                            66:'ind_m3',
                            67:'ind_m2',
                            68:'ind_m1',
                            69:'fecha_primer_gestion',
                            70:'telefono_mejor_gestion',
                            71:'fecha_gestion',
                            72:'contactabilidad',
                            73:'indicador_hoy',
                            74:'repeticion',
                            75:'llamadas',
                            76:'sms',
                            77:'correos',
                            78:'gescall',
                            79:'whatsapp',
                            80:'visitas',
                            81:'no_contacto',
                            82:'telefono_positivo',
                            83:'fec_ultima_marcacion',
                            84:'lista_robinson'})


    fn = pd.merge(df,inf,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infR,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infC,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    fn = pd.merge(fn,infD,on = ["deudor_id"]\
                ,how = "left",indicator = False)
    
    if 'infP' in locals():
    #    Cruce pagos
        fn = pd.merge(fn,infP,on = ["obligacion_id"]\
                    ,how = "left",indicator = False)
    #   ordenamiento
        lt = fn.columns.tolist()
        lt = lt[:27] + lt[(infP.shape[1]-1)*-1:] + lt[27:fn.shape[1]-(infP.shape[1]-1)]
        fn = fn[lt]

    return csv_o(fn,tablename)